#import pandas as pd
import glob
import fitz #part of pymupdf library -- use as pdf converter into txt format 
from openpyxl import load_workbook


files = glob.glob('*.pdf')
print(files)

wb = load_workbook('report.xlsx')

sheet = wb.active

attribute_list = ["ID", "cholesterol_v_", "hdl_cholesterol_v_", 
                  "triglycerides_v_", "ldl_cholesterol_v_", 
                  "chol_hdlc_ratio_v_", "non_hdl_cholesterol_v_",
                  "glucose_v_", "c-reactive_protein_v_", "insulin_v_"]

for i in range(1,4): # we are adding the columns upto 3 visit.
    if i == 1:
        sheet.cell(row=1, column=1).value = attribute_list[0]
    
    max_col_val = sheet.max_column
    for j in range(1,len(attribute_list)): 
        sheet.cell(row = 1, column = max_col_val+j).value = attribute_list[j]+str(i)
                
for file in files:
    with fitz.open(file) as doc:
        pymupdf_text = "" 
        for page in doc:
            pymupdf_text += page.get_text()
            
    text_list = pymupdf_text.split("\n")
# it will convert all txt data into which is in a single str into a single list.
    #print(text_list)
   
    remove_space = [x.strip(' ') for x in text_list]
    #print(remove_space)
    
    final_list = [ele.strip() for ele in text_list if ele.strip()]
    #print(final_list)
    
    attributes = ["CHOLESTEROL, TOTAL", "HDL CHOLESTEROL", "TRIGLYCERIDES",
                  "LDL-CHOLESTEROL","CHOL/HDLC RATIO", "NON HDL CHOLESTEROL",
                  "GLUCOSE", "C-REACTIVE PROTEIN", "INSULIN"]
    
    temp_dict = {}
    for data in attributes:
        temp_dict[data] = final_list[final_list.index(data)+1]
        
    cholesterol = temp_dict["CHOLESTEROL, TOTAL"].split(" ")[0]
    hdl_cholesterol = temp_dict["HDL CHOLESTEROL"].split(" ")[0]
    triglycerides = temp_dict["TRIGLYCERIDES"].split(" ")[0]
    ldl_cholesterol = temp_dict["LDL-CHOLESTEROL"].split(" ")[0]
    chol_hdlc_ratio = temp_dict["CHOL/HDLC RATIO"].split(" ")[0]
    non_hdl_cholesterol = temp_dict["NON HDL CHOLESTEROL"].split(" ")[0]
    glucose = temp_dict["GLUCOSE"].split(" ")[0]
    c_reactive_protein = temp_dict["C-REACTIVE PROTEIN"].split(" ")[0]
    insulin = temp_dict["INSULIN"].split(" ")[0]


    def isDigit(x):
        try:
            float(x)
            return True
        except ValueError:
            return False

    def isPresent(x, checklist):
        try:
            return checklist.index(x)
        except ValueError:
            return -1
         
    final_dict = {
        
        "ID": (final_list[final_list.index('Specimen:')-1].split(':')[1].strip())[0:-1],
        "cholesterol_v_": float(cholesterol) if isDigit(cholesterol) == True else " ",
        "hdl_cholesterol_v_": float(hdl_cholesterol) if isDigit(hdl_cholesterol) == True else " ",
        "triglycerides_v_": float(triglycerides) if isDigit(triglycerides) == True else " ",
        "ldl_cholesterol_v_": float(ldl_cholesterol) if isDigit(ldl_cholesterol) == True else " ",
        "chol_hdlc_ratio_v_": float(chol_hdlc_ratio) if isDigit(chol_hdlc_ratio) == True else " ",
        "non_hdl_cholesterol_v_": float(non_hdl_cholesterol) if isDigit(non_hdl_cholesterol) == True else " ",
        "glucose_v_": float(glucose) if isDigit(glucose) == True else " ",
        "c-reactive_protein_v_": float(c_reactive_protein) if isDigit(c_reactive_protein) == True else " ",
        "insulin_v_": float(insulin) if isDigit(insulin) == True else " "
        
        }
    
    keys = list(final_dict.keys())
    values = list(final_dict.values())
    
    Id = (final_list[final_list.index('Specimen:')-1].split(':')[1].strip())[0:-1]
    visit_num = (final_list[final_list.index('Specimen:')-1].split(':')[1].strip())[-1]    

    print(final_list[final_list.index('Specimen:')-1].split(':')[1].strip())
    
    m_row =sheet.max_row # Calculating how many time's patient tring to visted.
    Id_list = [sheet.cell (row=i, column=1).value for i in range(2, m_row + 1)]
    
    max_col = sheet.max_column
    columns = [sheet.cell(row=1, column=i).value for i in range(1, max_col + 1)]

    Id_index = isPresent(Id, Id_list) # Checking the perticular column is present or not.
    column_num = isPresent("cholesterol_v_" + visit_num, columns)
    
    
    if Id_index == -1:  # If the patient is visiting first time 
        sheet.cell(row = m_row + 1, column = 1).value = values[0]
        for i in range(1, len(values)):
            sheet.cell(row = m_row + 1, column= column_num+i).value = values[i]
    else:
        row_num = Id_index + 2
            
        if column_num == -1:
            for i in range(1, len(keys)):
                sheet.cell(row=1, column=max_col +i).value = keys[i]+visit_num
                
                sheet.cell(row=row_num, column=max_col+i).value = values[i]
        else:
            for i in range(1, len(values)):
                sheet.cell(row=row_num, column=column_num + i).value = values[i]

    wb.save('report.xlsx') 


  

 






























